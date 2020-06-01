import numpy as np
import pandas as pd
import torch
import torchvision
import torchvision.transforms as transforms
from torch.utils.data.sampler import SubsetRandomSampler
from .util import readmts_uci_har, transform_labels
def getDataset(dataset):
    transform = transforms.Compose([
        transforms.Resize((32, 32)),
        transforms.RandomHorizontalFlip(),
        transforms.RandomRotation(10),
        transforms.ToTensor(),
        ])

    if(dataset == 'CIFAR10'):
        trainset = torchvision.datasets.CIFAR10(root='./data', train=True, download=True, transform=transform)
        testset = torchvision.datasets.CIFAR10(root='./data', train=False, download=True, transform=transform)
        num_classes = 10
        inputs=3

    elif(dataset == 'CIFAR100'):
        trainset = torchvision.datasets.CIFAR100(root='./data', train=True, download=True, transform=transform)
        testset = torchvision.datasets.CIFAR100(root='./data', train=False, download=True, transform=transform)
        num_classes = 100
        inputs = 3
        
    elif(dataset == 'MNIST'):
        trainset = torchvision.datasets.MNIST(root='./data', train=True, download=True, transform=transform)
        testset = torchvision.datasets.MNIST(root='./data', train=False, download=True, transform=transform)
        num_classes = 10
        inputs = 1

    elif (dataset == 'UCI'):
        file_name = 'datasets/UCI_HAR_Dataset'
        x_train, y_train, x_test, y_test = readmts_uci_har(file_name)
        data = np.concatenate((x_train, x_test), axis=0)
        label = np.concatenate((y_train, y_test), axis=0)
        N = data.shape[0]
        ind = int(N * 0.9)
        x_train = data[:ind]
        y_train = label[:ind]
        x_test = data[ind:]
        y_test = label[ind:]
        y_train, y_test = transform_labels(y_train, y_test)

        x_train = x_train.swapaxes(1, 2)
        x_test = x_test.swapaxes(1, 2)

        x_train = torch.from_numpy(x_train).float()
        y_train = torch.from_numpy(y_train)
        print(x_train.size(), y_train.size())
        trainset = torch.utils.data.TensorDataset(x_train, y_train)

        x_test = torch.from_numpy(x_test).float()
        y_test = torch.from_numpy(y_test)
        print(x_test.size(), y_test.size())
        testset = torch.utils.data.TensorDataset(x_test, y_test)

        num_classes = 6

        inputs = 9

    return trainset, testset, inputs, num_classes

def getDataset_regression(dataset):
    if(dataset == 'uci_har'):
        file_name = 'datasets/UCI_HAR_Dataset'
        x_train, y_train, x_test, y_test = readmts_uci_har(file_name)
        data = np.concatenate((x_train, x_test), axis=0)
        label = np.concatenate((y_train, y_test), axis=0)
        N = data.shape[0]
        ind = int(N * 0.9)
        x_train = data[:ind]
        y_train = label[:ind]
        x_test = data[ind:]
        y_test = label[ind:]

        x_train = x_train.swapaxes(1, 2)
        x_test = x_test.swapaxes(1, 2)

        x_train = torch.from_numpy(x_train).float()
        y_train = torch.from_numpy(y_train).float()
        print(x_train.size(), y_train.size())
        trainset = torch.utils.data.TensorDataset(x_train, y_train)

        x_test = torch.from_numpy(x_test).float()
        y_test = torch.from_numpy(y_test).float()
        print(x_test.size(), y_test.size())
        testset = torch.utils.data.TensorDataset(x_test, y_test)

        inputs = 9
        outputs = 1

    elif(dataset == 'ccpp'):
        from openpyxl import load_workbook
        workbook = load_workbook(filename='datasets/CCPP/Folds5x2_pp.xlsx')
        sheet = workbook.get_sheet_by_name("Sheet1")
        data = []
        row_num = 2
        while row_num <= 9569:
            sample = []
            for i in range(5):
                sample.append(sheet.cell(row=row_num, column=i+1).value)
            sample = np.array(sample)
            data.append(sample)
            row_num = row_num + 1
        data = np.array(data)

        N = data.shape[0]
        ind = int(N * 0.9)
        train_data = data[:ind]
        test_data = data[ind:]
        x_train = train_data[:,:4]
        y_train = train_data[:,4]
        x_test = test_data[:,:4]
        y_test = test_data[:,4]

        x_means, x_stds = x_train.mean(axis=0), x_train.var(axis=0) ** 0.5
        y_means, y_stds = y_train.mean(axis=0), y_train.var(axis=0) ** 0.5

        x_train = (x_train - x_means) / x_stds
        y_train = (y_train - y_means) / y_stds

        x_test = (x_test - x_means) / x_stds
        y_test = (y_test - y_means) / y_stds

        x_train = torch.from_numpy(x_train).float()
        y_train = torch.from_numpy(y_train).float()
        print(x_train.size(), y_train.size())
        trainset = torch.utils.data.TensorDataset(x_train, y_train)

        x_test = torch.from_numpy(x_test).float()
        y_test = torch.from_numpy(y_test).float()
        print(x_test.size(), y_test.size())
        testset = torch.utils.data.TensorDataset(x_test, y_test)

        inputs = 4
        outputs = 1

    elif (dataset == 'ppg'):

        datafinal = pd.read_csv('datasets/PPG/data.csv', encoding='latin1', sep=',', index_col=0)
        datafinal.reset_index()
        d = datafinal
        a = list(d.columns)
        a.remove('label')
        list(d.columns)
        target = 'label'
        # d['Gender'] = d['Gender'].apply(lambda r: True if (r == 'f') else False)
        d['Gender'] = d['Gender'].apply(lambda r: 1 if (r == 'f') else 0)
        X = d[a]
        Y = d[target]

        x = np.array(X)
        y = np.array(Y)

        from sklearn.model_selection import train_test_split
        from sklearn.preprocessing import StandardScaler

        X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=27)

        sc = StandardScaler()
        X_train = sc.fit_transform(X_train)
        X_test = sc.transform(X_test)

        # N = len(y_train)
        # ind = int(N * 0.9)
        # X_val = X_train[ind:,:]
        # y_val = y_train[ind:]
        # X_train = X_train[:ind,:]
        # y_train = y_train[:ind]

        x_train = torch.from_numpy(X_train).float()
        y_train = torch.from_numpy(y_train).float()
        print(x_train.size(), y_train.size())
        trainset = torch.utils.data.TensorDataset(x_train, y_train)

        x_test = torch.from_numpy(X_test).float()
        y_test = torch.from_numpy(y_test).float()
        print(x_test.size(), y_test.size())
        testset = torch.utils.data.TensorDataset(x_test, y_test)

        inputs = 20
        outputs = 1
    elif(dataset == 'ppg2'):

        data = pd.read_csv('datasets/PPG/data.csv', encoding='latin1', sep=',')
        data_new = data.drop(['WEIGHT', 'Gender', 'AGE', 'HEIGHT', 'SKIN', 'SPORT', 'Activity', 'EMG', 'EDA', 'Temp'],
                             axis=1)
        data_new = np.array(data_new)

        label = data_new[::8, 0]
        y = np.delete(label, -1, axis=0)
        n = len(y)

        data = []
        for i in range(n):
            ts = data_new[i * 8:(i + 1) * 8, 1:]
            data.append(ts)

        x = np.array(data)

        from sklearn.model_selection import train_test_split
        from sklearn.preprocessing import StandardScaler

        X_train, X_test, y_train, y_test = train_test_split(x, y, test_size=0.1, random_state=27)

        X_train = X_train.swapaxes(1, 2)
        X_test = X_test.swapaxes(1, 2)

        x_train = torch.from_numpy(X_train).float()
        y_train = torch.from_numpy(y_train).float()
        print(x_train.size(), y_train.size())
        trainset = torch.utils.data.TensorDataset(x_train, y_train)

        x_test = torch.from_numpy(X_test).float()
        y_test = torch.from_numpy(y_test).float()
        print(x_test.size(), y_test.size())
        testset = torch.utils.data.TensorDataset(x_test, y_test)

        inputs = 11
        outputs = 1

    return trainset, testset, inputs, outputs


def getDataloader(trainset, testset, valid_size, batch_size, num_workers):
    num_train = len(trainset)
    indices = list(range(num_train))
    np.random.shuffle(indices)
    split = int(np.floor(valid_size * num_train))
    train_idx, valid_idx = indices[split:], indices[:split]

    train_sampler = SubsetRandomSampler(train_idx)
    valid_sampler = SubsetRandomSampler(valid_idx)

    train_loader = torch.utils.data.DataLoader(trainset, batch_size=batch_size,
        sampler=train_sampler, num_workers=num_workers)
    print(len(train_loader))
    valid_loader = torch.utils.data.DataLoader(trainset, batch_size=batch_size, 
        sampler=valid_sampler, num_workers=num_workers)
    print(len(valid_loader))
    test_loader = torch.utils.data.DataLoader(testset, batch_size=batch_size, 
        num_workers=num_workers)
    print(len(test_loader))

    return train_loader, valid_loader, test_loader